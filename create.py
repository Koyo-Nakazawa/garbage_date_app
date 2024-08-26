import openpyxl.workbook
from config import Areas, GarbageTyeps, CollectionTypes
import datetime
import openpyxl


# 地区テーブルのデータ作成
def create_area(area_name, collection_type_id):
    Areas.create(area_name=area_name, collection_type_id=collection_type_id)


# ごみ種別テーブルのデータ作成
def create_garbage_type(garbage_type_name):
    GarbageTyeps.create(garbage_type_name=garbage_type_name)


# 収集種別テーブルのデータ作成
def create_collection_type(collection_type_id, garbage_type_id, collection_date):
    CollectionTypes.create(
        collection_type_id=collection_type_id,
        garbage_type_id=garbage_type_id,
        collection_date=collection_date,
    )


if __name__ == "__main__":
    # とりあえず一件ずつ作成する処理は完了
    # CSVを読み込んで一括で作成する処理をあとで追加すること

    # Excelからダミーデータを読み込む
    wb = openpyxl.load_workbook("テーブル定義.xlsx")

    # 地区テーブルの読み込み
    ws = wb["areas"]
    cnt = 2
    while True:
        area_name = ws.cell(cnt, 2).value
        collection_type_id = ws.cell(cnt, 3).value
        if area_name is None:
            break
        create_area(area_name, collection_type_id)
        cnt += 1

    # ごみ種別テーブルの読み込み
    ws = wb["garbage_types"]
    cnt = 2
    while True:
        garbage_type_name = ws.cell(cnt, 2).value
        if garbage_type_name is None:
            break
        create_garbage_type(garbage_type_name)
        cnt += 1

    # 収集種別テーブルの読み込み
    ws = wb["collection_types"]
    cnt = 2
    while True:
        collection_type_id = ws.cell(cnt, 1).value
        garbage_type_id = ws.cell(cnt, 2).value
        collection_date = ws.cell(cnt, 3).value
        if collection_type_id is None:
            break
        create_collection_type(collection_type_id, garbage_type_id, collection_date)
        cnt += 1

    # create_area("箱清水", 1)
    # create_garbage_type("可燃ごみ")
    # create_collection_type(1, 1, datetime.date(2024, 8, 26))
